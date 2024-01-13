import openpyxl as xl

wb = xl.load_workbook("suparco.xlsx")
ws = wb.active
row_num  = 1
# col_num  = 2
# cell = ""
# if row_num == 1:
#     row_num = "A"
#     cell = row_num + str(col_num)
# print(cell)
for row in ws.iter_rows(values_only=True):

    print(row_num)
    print(row)
    row_num = row_num + 1
#     ws.cell(row = row_num, column = 2, value = "12345abc")
#     # ws["A1"]

# wb.save("suparco.xlsx")



# ws.delete_rows()
# # row_num = 0
# # for row in ws.iter_rows(values_only=True):
# #     print(row_num)
# #     print(row)
# #     row_num=  row_num + 1