import openpyxl as xl
# workbook=xl.Workbook()
# workbook.save("suparco.xlsx")
# workbook_already_made=xl.load_workbook(filename="suparco.xlsx")
#
# sheet1=workbook_already_made.active
# sheet1["A5"]=7425
# sheet1["C98"]=24253
# sheet1.append([34,1335,"beaconhouse"])
# sheet1.append([53,123,"suparco is boring"])
# mylist=[]
# mylist.extend([43,56,214,54])
# mylist.extend([24,656,214,65])
# mylist.extend([43,56,636,65])
# mylist.extend([43,6456.8,214,741])
# mylist.extend([43,248,2,65])
# sheet1.append(mylist)
def add_product():
    workbook_already_made = xl.load_workbook(filename="suparco.xlsx")

    sheet1 = workbook_already_made.active

    number_of_products=int(input("please tell how many products to be putin system"))
    list_of_products=[]
    for i in range(1,number_of_products):
        name=input(("please add your product name #",i))
        quantity=int(input(("please add product quantity",i)))
        code=int(input(("please add product code",i)))
        price=int(input(("please add product price",i)))
        sheet1.append([name,quantity,code,price])
        print("product added successfully")


    workbook_already_made.save("suparco.xlsx")


    # product=[name,quantity,code,price]
    # return product
# product_made=add_product()
# # print(product_made)
# sheet1.append(product_made)

add_product()
# for i in list_of_products_made:
#     sheet1.append(i)










